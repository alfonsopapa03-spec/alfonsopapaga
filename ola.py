"""
Sistema de Registro y Legalización de Anticipos - Transporte de Carga
Colombia - Conectado a Supabase (PostgreSQL)
v15: 
  - Fechas mostradas en formato DD/MM/YYYY
  - Períodos con vacaciones tomadas pero SIN pago en dinero → 🟡 advertencia (no verde)
  - Pago en dinero registrado → 💵 info (no marca como completado; indica que aún faltan vacaciones físicas)
"""

import streamlit as st
import psycopg2
from psycopg2 import pool
import pandas as pd
from datetime import datetime, timedelta, date
from math import floor
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==================== CONFIGURACIÓN ====================
SUPABASE_DB_URL = "postgresql://postgres.ntnpckmbyfmjhfskfwyu:Conejito100#@aws-1-us-east-1.pooler.supabase.com:6543/postgres"
DIAS_VACACIONES_ANUALES = 15

# ==================== FORMATO COLOMBIANO ====================
def fmt(valor):
    if valor is None:
        return "0"
    try:
        return f"{int(float(valor)):,}".replace(',', '.')
    except:
        return str(valor)

def limpiar(texto):
    if not texto:
        return 0.0
    try:
        return float(str(texto).replace('.', '').replace(',', '.'))
    except:
        return 0.0

def hora_colombia():
    return datetime.utcnow() - timedelta(hours=5)

def fmt_fecha(valor):
    """Formatea una fecha o string de fecha a DD/MM/YYYY."""
    if valor is None:
        return "—"
    try:
        if isinstance(valor, (date, datetime)):
            return valor.strftime('%d/%m/%Y')
        return pd.to_datetime(valor).strftime('%d/%m/%Y')
    except:
        return str(valor)[:10]

# ==================== CONNECTION POOL (singleton) ====================
@st.cache_resource
def get_pool():
    return psycopg2.pool.ThreadedConnectionPool(
        minconn=1,
        maxconn=5,
        dsn=SUPABASE_DB_URL,
        connect_timeout=10,
        options="-c statement_timeout=15000"
    )

def get_conn():
    return get_pool().getconn()

def put_conn(conn):
    get_pool().putconn(conn)

# ==================== ALERTAS ANTICIPOS ====================
def clasificar_alerta(fecha_viaje):
    hoy = hora_colombia().date()
    try:
        fv = fecha_viaje.date() if hasattr(fecha_viaje, 'date') else pd.to_datetime(fecha_viaje).date()
    except:
        return 0, "ok"
    dias = (hoy - fv).days
    if dias <= 3:
        return dias, "ok"
    elif dias <= 7:
        return dias, "warning"
    else:
        return dias, "critical"

def badge_alerta(dias, nivel):
    if nivel == "critical":
        return f"🔴 {dias}d"
    elif nivel == "warning":
        return f"🟡 {dias}d"
    else:
        return f"🟢 {dias}d"

# ==================== LÓGICA VACACIONES v15 ====================

def calcular_vacaciones(conductor: str, fecha_ingreso: date, df_vac: pd.DataFrame, hoy: date) -> dict:
    anios_completos = 0
    periodos = []

    n = 1
    while True:
        try:
            inicio_periodo = fecha_ingreso.replace(year=fecha_ingreso.year + (n - 1))
        except ValueError:
            inicio_periodo = date(fecha_ingreso.year + (n - 1), fecha_ingreso.month, 28)
        try:
            fin_periodo = fecha_ingreso.replace(year=fecha_ingreso.year + n)
        except ValueError:
            fin_periodo = date(fecha_ingreso.year + n, fecha_ingreso.month, 28)

        if fin_periodo > hoy:
            proxima_fecha = fin_periodo
            break

        periodos.append({
            "anio": n,
            "inicio": inicio_periodo,
            "fin": fin_periodo,
            "label": f"Año {n}  ({inicio_periodo.strftime('%d/%m/%Y')} → {fin_periodo.strftime('%d/%m/%Y')})",
        })
        anios_completos = n
        n += 1

    if anios_completos == 0:
        try:
            proxima_fecha = fecha_ingreso.replace(year=fecha_ingreso.year + 1)
        except ValueError:
            proxima_fecha = date(fecha_ingreso.year + 1, fecha_ingreso.month, 28)

    dias_generados = anios_completos * DIAS_VACACIONES_ANUALES

    vac_cond = df_vac[df_vac["conductor"] == conductor] if not df_vac.empty else pd.DataFrame()
    dias_usados = int(vac_cond["dias"].sum()) if not vac_cond.empty else 0
    dias_disponibles = dias_generados - dias_usados
    dias_vencidos = max(0, dias_disponibles)

    dias_para_proxima = (proxima_fecha - hoy).days

    return {
        "anios_completos": anios_completos,
        "dias_generados": dias_generados,
        "dias_usados": dias_usados,
        "dias_disponibles": dias_disponibles,
        "dias_vencidos": dias_vencidos,
        "proxima_fecha": proxima_fecha,
        "dias_para_proxima": dias_para_proxima,
        "periodos": periodos,
        "registros": vac_cond.to_dict("records") if not vac_cond.empty else [],
    }


# ==================== EXPORTAR EXCEL ANTICIPOS ====================
def generar_excel(df: pd.DataFrame, titulo: str = "Anticipos") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Anticipos"
    color_header   = "1F4E79"
    color_critico  = "FCE4EC"
    color_warning  = "FFF9C4"
    color_ok       = "E8F5E9"
    color_leg      = "E3F2FD"
    color_subtotal = "BBDEFB"
    font_header = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    font_titulo = Font(name="Arial", bold=True, size=13, color="1F4E79")
    font_normal = Font(name="Arial", size=9)
    font_bold   = Font(name="Arial", bold=True, size=9)
    font_red    = Font(name="Arial", bold=True, size=9, color="C62828")
    font_subtot = Font(name="Arial", bold=True, size=10, color="1F4E79")
    thin   = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    ws.merge_cells("A1:M1")
    ws["A1"] = f"Reporte de Anticipos — {titulo}"
    ws["A1"].font = font_titulo
    ws["A1"].alignment = center
    ws.merge_cells("A2:M2")
    ws["A2"] = f"Generado: {hora_colombia().strftime('%d/%m/%Y %H:%M')} (hora Colombia)"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="757575")
    ws["A2"].alignment = center
    columnas = ["ID","Manifiesto","Fecha viaje","Placa","Conductor","Cliente","Origen","Destino",
                "Anticipo (COP)","Estado","Días pend.","Legalizado por","Fecha legalización"]
    row_header = 4
    for col_idx, col_name in enumerate(columnas, start=1):
        cell = ws.cell(row=row_header, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = PatternFill("solid", fgColor=color_header)
        cell.alignment = center
        cell.border = border
    for row_idx, (_, row) in enumerate(df.iterrows(), start=row_header + 1):
        legalizado = bool(row.get("legalizado", False))
        dias, nivel = clasificar_alerta(row.get("fecha_viaje"))
        if legalizado:
            row_color = color_leg
        elif nivel == "critical":
            row_color = color_critico
        elif nivel == "warning":
            row_color = color_warning
        else:
            row_color = color_ok
        fill = PatternFill("solid", fgColor=row_color)
        valores = [
            row.get("id",""), row.get("manifiesto",""),
            str(row.get("fecha_viaje",""))[:10], row.get("placa",""),
            row.get("conductor",""), row.get("cliente",""),
            row.get("origen",""), row.get("destino",""),
            int(row.get("valor_anticipo",0)),
            "Legalizado" if legalizado else "Pendiente",
            "" if legalizado else dias,
            row.get("legalizado_por","") or "",
            str(row.get("fecha_legalizacion","") or "")[:16],
        ]
        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = fill; cell.border = border
            cell.alignment = center if col_idx in [1,10,11] else left
            if col_idx == 9 and not legalizado and nivel == "critical":
                cell.font = font_red
            elif col_idx == 9:
                cell.font = font_bold
            else:
                cell.font = font_normal
    total_row = row_header + len(df) + 2
    ws.cell(row=total_row, column=8, value="TOTAL ANTICIPOS:").font = font_subtot
    ws.cell(row=total_row, column=8).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=9, value=f'=SUM(I{row_header+1}:I{row_header+len(df)})').font = font_subtot
    ws.cell(row=total_row, column=9).fill = PatternFill("solid", fgColor=color_subtotal)
    ws.cell(row=total_row, column=9).border = border
    ws.cell(row=total_row, column=9).alignment = center
    anchos = [6,14,13,10,22,20,18,18,18,16,10,22,20]
    for col_idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho
    for row_idx in range(row_header+1, row_header+len(df)+1):
        ws.cell(row=row_idx, column=9).number_format = '#,##0'
    ws.freeze_panes = f"A{row_header + 1}"
    output = BytesIO(); wb.save(output); output.seek(0)
    return output


# ==================== EXPORTAR EXCEL PRÉSTAMOS ====================
def generar_excel_prestamos(df_prestamos: pd.DataFrame, df_pagos: pd.DataFrame) -> BytesIO:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Préstamos"
    color_h    = "1F4E79"
    color_paz  = "E8F5E9"
    color_deu  = "FCE4EC"
    fh  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fn  = Font(name="Arial", size=9)
    ft  = Font(name="Arial", bold=True, size=13, color="1F4E79")
    thin   = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left_a = Alignment(horizontal="left",   vertical="center")
    ws1.merge_cells("A1:I1")
    ws1["A1"] = f"Reporte de Préstamos — {hora_colombia().strftime('%d/%m/%Y %H:%M')}"
    ws1["A1"].font = ft; ws1["A1"].alignment = center
    cols_prest = ["ID","Conductor","Fecha préstamo","Monto total","Total pagado","Saldo pendiente","Estado","Motivo","Observaciones"]
    rh = 3
    for ci, cn in enumerate(cols_prest, 1):
        cell = ws1.cell(row=rh, column=ci, value=cn)
        cell.font = fh
        cell.fill = PatternFill("solid", fgColor=color_h)
        cell.alignment = center; cell.border = border
    for ri, (_, row) in enumerate(df_prestamos.iterrows(), start=rh+1):
        paz = row.get("estado","") == "saldado"
        fill = PatternFill("solid", fgColor=color_paz if paz else color_deu)
        pid = row.get("id",0)
        pagos_conductor = df_pagos[df_pagos["prestamo_id"]==pid]["monto_pago"].sum() if not df_pagos.empty else 0
        saldo = max(0, int(row.get("monto_total",0)) - int(pagos_conductor))
        valores = [
            pid, row.get("conductor",""),
            str(row.get("fecha_prestamo",""))[:10],
            int(row.get("monto_total",0)),
            int(pagos_conductor), saldo,
            "Paz y salvo" if paz else "Pendiente",
            row.get("motivo","") or "", row.get("observaciones","") or ""
        ]
        for ci, val in enumerate(valores, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.border = border; cell.font = fn
            cell.alignment = center if ci in [1,4,5,6,7] else left_a
            if ci in [4,5,6]: cell.number_format = '#,##0'
    anchos1 = [6,22,14,16,16,16,14,22,28]
    for ci, aw in enumerate(anchos1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = aw
    ws1.freeze_panes = f"A{rh+1}"

    if not df_pagos.empty:
        ws2 = wb.create_sheet("Detalle pagos")
        ws2.merge_cells("A1:F1")
        ws2["A1"] = "Detalle de pagos / descuentos"
        ws2["A1"].font = ft; ws2["A1"].alignment = center
        cols_pago = ["ID pago","Préstamo ID","Conductor","Fecha pago","Monto descuento","Observaciones"]
        rh2 = 3
        for ci, cn in enumerate(cols_pago, 1):
            cell = ws2.cell(row=rh2, column=ci, value=cn)
            cell.font = fh
            cell.fill = PatternFill("solid", fgColor=color_h)
            cell.alignment = center; cell.border = border
        for ri, (_, row) in enumerate(df_pagos.iterrows(), start=rh2+1):
            fill2 = PatternFill("solid", fgColor="F3F3F3")
            pid2 = row.get("prestamo_id",0)
            cond2 = ""
            if not df_prestamos.empty:
                match = df_prestamos[df_prestamos["id"]==pid2]
                if not match.empty:
                    cond2 = match.iloc[0].get("conductor","")
            valores2 = [
                row.get("id",""), pid2, cond2,
                str(row.get("fecha_pago",""))[:10],
                int(row.get("monto_pago",0)),
                row.get("observaciones","") or ""
            ]
            for ci, val in enumerate(valores2, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.fill = fill2; cell.border = border; cell.font = fn
                cell.alignment = center if ci in [1,2,4,5] else left_a
                if ci == 5: cell.number_format = '#,##0'
        anchos2 = [8,12,22,14,18,30]
        for ci, aw in enumerate(anchos2, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = aw
        ws2.freeze_panes = f"A{rh2+1}"

    output = BytesIO(); wb.save(output); output.seek(0)
    return output


# ==================== EXPORTAR EXCEL VACACIONES v15 ====================
def generar_excel_vacaciones(df_info: pd.DataFrame, df_vac: pd.DataFrame, df_pagos_vac: pd.DataFrame, conductores: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen vacaciones"
    color_h    = "1F4E79"
    color_venc = "FCE4EC"
    color_ok_  = "E8F5E9"
    color_pag  = "E3F2FD"
    fh = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fn = Font(name="Arial", size=9)
    ft = Font(name="Arial", bold=True, size=13, color="1F4E79")
    thin   = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left_a = Alignment(horizontal="left",   vertical="center")
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Reporte de Vacaciones — {hora_colombia().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = ft; ws["A1"].alignment = center

    cols = [
        "Conductor", "Fecha ingreso", "Años completos",
        "Días generados (×15)", "Días usados",
        "Días pendientes", "Estado",
        "Próxima vacación", "Días para próxima"
    ]
    rh = 3
    for ci, cn in enumerate(cols, 1):
        cell = ws.cell(row=rh, column=ci, value=cn)
        cell.font = fh
        cell.fill = PatternFill("solid", fgColor=color_h)
        cell.alignment = center; cell.border = border

    hoy = hora_colombia().date()
    for ri, cond in enumerate(sorted(conductores), start=rh+1):
        info_row = df_info[df_info["conductor"] == cond].iloc[0] \
            if not df_info.empty and (df_info["conductor"] == cond).any() else None

        if info_row is not None and info_row.get("fecha_ingreso") is not None:
            fi = pd.to_datetime(info_row["fecha_ingreso"]).date()
            calc = calcular_vacaciones(cond, fi, df_vac, hoy)

            pagos_cond = df_pagos_vac[df_pagos_vac["conductor"] == cond] if not df_pagos_vac.empty else pd.DataFrame()
            anios_pagados = set(pagos_cond["anio_laboral"].tolist()) if not pagos_cond.empty else set()
            periodos_pendientes_sin_pago = sum(
                1 for p in calc["periodos"]
                if p["anio"] not in anios_pagados
                and (DIAS_VACACIONES_ANUALES - sum(
                    int(r.get("dias", 0)) for r in calc["registros"]
                    if r.get("fecha_inicio") is not None and (
                        pd.to_datetime(r["fecha_inicio"]).date() >= p["inicio"]
                        and pd.to_datetime(r["fecha_inicio"]).date() < p["fin"]
                    )
                )) > 0
            )

            if periodos_pendientes_sin_pago > 0:
                estado_txt = "Con días pendientes"
                fill_color = color_venc
            elif calc["anios_completos"] > 0:
                estado_txt = "Al día / Pagado"
                fill_color = color_ok_
            else:
                estado_txt = "Sin períodos aún"
                fill_color = color_ok_

            valores = [
                cond, str(fi),
                calc["anios_completos"],
                calc["dias_generados"],
                calc["dias_usados"],
                calc["dias_vencidos"],
                estado_txt,
                str(calc["proxima_fecha"]),
                calc["dias_para_proxima"],
            ]
        else:
            fill_color = "F3F3F3"
            valores = [cond, "—", "—", "—", "—", "—", "Sin fecha ingreso", "—", "—"]

        fill = PatternFill("solid", fgColor=fill_color)
        for ci, val in enumerate(valores, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill; cell.border = border; cell.font = fn
            cell.alignment = center if ci in [3,4,5,6,9] else left_a

    anchos = [28, 14, 16, 18, 14, 16, 20, 16, 18]
    for ci, aw in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(ci)].width = aw
    ws.freeze_panes = f"A{rh+1}"

    ws2 = wb.create_sheet("Historial tomadas")
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Historial de vacaciones tomadas por conductor"
    ws2["A1"].font = ft; ws2["A1"].alignment = center
    cols2 = ["Conductor", "Fecha inicio", "Fecha fin", "Días", "Observaciones", "Registrado por", "Fecha registro"]
    for ci, cn in enumerate(cols2, 1):
        cell = ws2.cell(row=3, column=ci, value=cn)
        cell.font = fh; cell.fill = PatternFill("solid", fgColor=color_h)
        cell.alignment = center; cell.border = border

    ri2 = 4
    if not df_vac.empty:
        for _, row in df_vac.sort_values(["conductor", "fecha_inicio"]).iterrows():
            fill2 = PatternFill("solid", fgColor=color_ok_)
            vals2 = [
                row.get("conductor",""),
                str(row.get("fecha_inicio",""))[:10],
                str(row.get("fecha_fin",""))[:10],
                row.get("dias",""),
                row.get("observaciones","") or "",
                row.get("registrado_por","") or "",
                str(row.get("fecha_registro",""))[:16],
            ]
            for ci, val in enumerate(vals2, 1):
                cell = ws2.cell(row=ri2, column=ci, value=val)
                cell.fill = fill2; cell.border = border; cell.font = fn
                cell.alignment = center if ci in [4] else left_a
            ri2 += 1

    anchos2 = [28, 14, 12, 8, 28, 20, 18]
    for ci, aw in enumerate(anchos2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = aw
    ws2.freeze_panes = "A4"

    ws3 = wb.create_sheet("Pagos en dinero")
    ws3.merge_cells("A1:H1")
    ws3["A1"] = "Pagos de vacaciones en dinero por período anual"
    ws3["A1"].font = ft; ws3["A1"].alignment = center
    cols3 = ["Conductor", "Año laboral", "Período", "Monto pagado (COP)", "Fecha pago", "Registrado por", "Observaciones", "Fecha registro"]
    for ci, cn in enumerate(cols3, 1):
        cell = ws3.cell(row=3, column=ci, value=cn)
        cell.font = fh; cell.fill = PatternFill("solid", fgColor=color_h)
        cell.alignment = center; cell.border = border

    ri3 = 4
    if not df_pagos_vac.empty:
        for _, row in df_pagos_vac.sort_values(["conductor", "anio_laboral"]).iterrows():
            fill3 = PatternFill("solid", fgColor=color_pag)
            vals3 = [
                row.get("conductor",""),
                row.get("anio_laboral",""),
                row.get("periodo_label","") or "",
                int(row.get("monto_cop", 0)),
                str(row.get("fecha_pago",""))[:10],
                row.get("registrado_por","") or "",
                row.get("observaciones","") or "",
                str(row.get("fecha_registro",""))[:16],
            ]
            for ci, val in enumerate(vals3, 1):
                cell = ws3.cell(row=ri3, column=ci, value=val)
                cell.fill = fill3; cell.border = border; cell.font = fn
                cell.alignment = center if ci in [2, 4] else left_a
                if ci == 4: cell.number_format = '#,##0'
            ri3 += 1

    anchos3 = [28, 12, 32, 20, 14, 20, 28, 18]
    for ci, aw in enumerate(anchos3, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = aw
    ws3.freeze_panes = "A4"

    output = BytesIO(); wb.save(output); output.seek(0)
    return output


# ==================== PLACAS ====================
PLACAS = [
    "NOX459","NOX460","NOX461","SON047","SON048",
    "SOP148","SOP149","SOP150","SRO661","SRO672",
    "TMW882","TRL282","TRL298","UYQ308","UYV084",
    "UYY788","PSX350"
]

CONDUCTORES_DEFAULT = [
    "CARLOS TAFUR","CHRISTIAN MARTINEZ","EDGAR DE JESUS",
    "EDUARDO OLIVARES","FLAVIO MALTE","GONZALO","ISAIAS VESGA",
    "JOSE ORTEGA","JULIAN CALETH","PEDRO JR","RAMON TAFUR",
    "REIMUR MANUEL","SLITH ORTEGA","YEIMI DUQUE","SIN CONDUCTOR ASIGNADO",
]

CLIENTES_DEFAULT = [
    "GLOBO EXPRESS","MOTOTRANSPORTAMOS","CARGO ANDINA","TRANSOLICAR","SUCLOGISTIC",
]


# ==================== BASE DE DATOS ====================
class DB:
    def _exec(self, query, params=None, fetch=None):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute(query, params)
            if fetch == "all":
                return cur.fetchall(), [d[0] for d in cur.description]
            elif fetch == "one":
                return cur.fetchone()
            else:
                conn.commit()
                return True
        except Exception as e:
            conn.rollback()
            st.error(f"DB error: {e}")
            return None
        finally:
            put_conn(conn)

    def _query_df(self, query, params=None):
        conn = get_conn()
        try:
            df = pd.read_sql_query(query, conn, params=params)
            return df
        except Exception as e:
            st.error(f"DB query error: {e}")
            return pd.DataFrame()
        finally:
            put_conn(conn)

    def init_tablas(self):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS anticipos_v1 (
                    id SERIAL PRIMARY KEY,
                    fecha_viaje DATE NOT NULL,
                    fecha_registro TIMESTAMP NOT NULL,
                    placa TEXT NOT NULL,
                    conductor TEXT NOT NULL,
                    cliente TEXT NOT NULL,
                    origen TEXT NOT NULL,
                    destino TEXT NOT NULL,
                    valor_anticipo BIGINT NOT NULL,
                    observaciones TEXT,
                    legalizado BOOLEAN DEFAULT FALSE,
                    fecha_legalizacion TIMESTAMP,
                    legalizado_por TEXT,
                    obs_legalizacion TEXT
                )
            """)
            cur.execute("ALTER TABLE anticipos_v1 ADD COLUMN IF NOT EXISTS manifiesto TEXT DEFAULT ''")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS clientes_extra (
                    id SERIAL PRIMARY KEY,
                    nombre TEXT UNIQUE NOT NULL,
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS conductores_extra (
                    id SERIAL PRIMARY KEY,
                    nombre TEXT UNIQUE NOT NULL,
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS conductores_info (
                    id SERIAL PRIMARY KEY,
                    conductor TEXT UNIQUE NOT NULL,
                    fecha_ingreso DATE,
                    observaciones TEXT,
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS vacaciones (
                    id SERIAL PRIMARY KEY,
                    conductor TEXT NOT NULL,
                    fecha_inicio DATE NOT NULL,
                    fecha_fin DATE NOT NULL,
                    dias INTEGER NOT NULL,
                    anio_laboral INTEGER,
                    observaciones TEXT,
                    registrado_por TEXT,
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            cur.execute("ALTER TABLE vacaciones ADD COLUMN IF NOT EXISTS anio_laboral INTEGER")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS prestamos (
                    id SERIAL PRIMARY KEY,
                    conductor TEXT NOT NULL,
                    monto_total BIGINT NOT NULL,
                    fecha_prestamo DATE NOT NULL,
                    motivo TEXT,
                    observaciones TEXT,
                    estado TEXT DEFAULT 'activo',
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS pagos_prestamos (
                    id SERIAL PRIMARY KEY,
                    prestamo_id INTEGER NOT NULL REFERENCES prestamos(id) ON DELETE CASCADE,
                    monto_pago BIGINT NOT NULL,
                    fecha_pago DATE NOT NULL,
                    observaciones TEXT,
                    registrado_por TEXT,
                    fecha_registro TIMESTAMP NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS vacaciones_pagos (
                    id SERIAL PRIMARY KEY,
                    conductor TEXT NOT NULL,
                    anio_laboral INTEGER NOT NULL,
                    periodo_label TEXT,
                    monto_cop BIGINT NOT NULL,
                    fecha_pago DATE NOT NULL,
                    observaciones TEXT,
                    registrado_por TEXT,
                    fecha_registro TIMESTAMP NOT NULL,
                    UNIQUE (conductor, anio_laboral)
                )
            """)
            conn.commit()
        except Exception as e:
            conn.rollback()
            st.error(f"Error inicializando tablas: {e}")
        finally:
            put_conn(conn)

    # ---- Clientes ----
    def obtener_clientes_extra(self):
        return self._query_df("SELECT * FROM clientes_extra ORDER BY nombre")

    def agregar_cliente(self, nombre):
        return bool(self._exec(
            "INSERT INTO clientes_extra (nombre, fecha_registro) VALUES (%s, %s)",
            (nombre.strip().upper(), hora_colombia())
        ))

    def eliminar_cliente(self, cliente_id):
        self._exec("DELETE FROM clientes_extra WHERE id = %s", (cliente_id,))

    # ---- Conductores extra ----
    def obtener_conductores_extra(self):
        return self._query_df("SELECT * FROM conductores_extra ORDER BY nombre")

    def agregar_conductor(self, nombre):
        return bool(self._exec(
            "INSERT INTO conductores_extra (nombre, fecha_registro) VALUES (%s, %s)",
            (nombre.strip().upper(), hora_colombia())
        ))

    def editar_conductor(self, conductor_id, nombre_nuevo):
        return bool(self._exec(
            "UPDATE conductores_extra SET nombre = %s WHERE id = %s",
            (nombre_nuevo.strip().upper(), conductor_id)
        ))

    def eliminar_conductor(self, conductor_id):
        self._exec("DELETE FROM conductores_extra WHERE id = %s", (conductor_id,))

    # ---- Conductores info ----
    def obtener_info_conductor(self, conductor):
        df = self._query_df("SELECT * FROM conductores_info WHERE conductor = %s", (conductor,))
        return df.iloc[0] if not df.empty else None

    def obtener_todos_info_conductores(self):
        return self._query_df("SELECT * FROM conductores_info ORDER BY conductor")

    def guardar_info_conductor(self, conductor, fecha_ingreso, observaciones=""):
        return bool(self._exec("""
            INSERT INTO conductores_info (conductor, fecha_ingreso, observaciones, fecha_registro)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (conductor) DO UPDATE
            SET fecha_ingreso = EXCLUDED.fecha_ingreso,
                observaciones = EXCLUDED.observaciones
        """, (conductor.strip().upper(), fecha_ingreso, observaciones.strip(), hora_colombia())))

    # ---- Vacaciones ----
    def obtener_vacaciones(self, conductor=None):
        if conductor:
            return self._query_df(
                "SELECT * FROM vacaciones WHERE conductor = %s ORDER BY fecha_inicio DESC",
                (conductor,)
            )
        return self._query_df("SELECT * FROM vacaciones ORDER BY fecha_inicio DESC")

    def registrar_vacacion(self, data):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO vacaciones (conductor, fecha_inicio, fecha_fin, dias, anio_laboral,
                    observaciones, registrado_por, fecha_registro)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
            """, (
                data['conductor'], data['fecha_inicio'], data['fecha_fin'],
                int(data['dias']), data.get('anio_laboral'),
                data.get('observaciones',''),
                data.get('registrado_por','').strip().upper(), hora_colombia()
            ))
            nuevo_id = cur.fetchone()[0]
            conn.commit()
            return nuevo_id
        except Exception as e:
            conn.rollback()
            st.error(f"Error registrando vacación: {e}")
            return None
        finally:
            put_conn(conn)

    def actualizar_vacacion(self, vac_id, data):
        return bool(self._exec("""
            UPDATE vacaciones
            SET fecha_inicio = %s, fecha_fin = %s, dias = %s, observaciones = %s
            WHERE id = %s
        """, (
            data['fecha_inicio'], data['fecha_fin'],
            int(data['dias']), data.get('observaciones',''), vac_id
        )))

    def eliminar_vacacion(self, vac_id):
        self._exec("DELETE FROM vacaciones WHERE id = %s", (vac_id,))

    # ---- Pagos de vacaciones en dinero (v15) ----
    def obtener_pagos_vacaciones(self, conductor=None):
        if conductor:
            return self._query_df(
                "SELECT * FROM vacaciones_pagos WHERE conductor = %s ORDER BY anio_laboral",
                (conductor,)
            )
        return self._query_df("SELECT * FROM vacaciones_pagos ORDER BY conductor, anio_laboral")

    def registrar_pago_vacacion(self, data):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO vacaciones_pagos
                    (conductor, anio_laboral, periodo_label, monto_cop, fecha_pago,
                     observaciones, registrado_por, fecha_registro)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (conductor, anio_laboral) DO UPDATE
                SET monto_cop       = EXCLUDED.monto_cop,
                    periodo_label   = EXCLUDED.periodo_label,
                    fecha_pago      = EXCLUDED.fecha_pago,
                    observaciones   = EXCLUDED.observaciones,
                    registrado_por  = EXCLUDED.registrado_por,
                    fecha_registro  = EXCLUDED.fecha_registro
                RETURNING id
            """, (
                data['conductor'],
                int(data['anio_laboral']),
                data.get('periodo_label', ''),
                int(data['monto_cop']),
                data['fecha_pago'],
                data.get('observaciones', '').strip(),
                data.get('registrado_por', '').strip().upper(),
                hora_colombia()
            ))
            nuevo_id = cur.fetchone()[0]
            conn.commit()
            return nuevo_id
        except Exception as e:
            conn.rollback()
            st.error(f"Error registrando pago de vacaciones: {e}")
            return None
        finally:
            put_conn(conn)

    def eliminar_pago_vacacion(self, pago_id):
        self._exec("DELETE FROM vacaciones_pagos WHERE id = %s", (pago_id,))

    # ---- Préstamos ----
    def obtener_prestamos(self, conductor=None, estado=None):
        q = "SELECT * FROM prestamos WHERE 1=1"
        params = []
        if conductor:
            q += " AND conductor = %s"; params.append(conductor)
        if estado and estado != "Todos":
            q += " AND estado = %s"; params.append(estado)
        q += " ORDER BY fecha_prestamo DESC, fecha_registro DESC"
        return self._query_df(q, params if params else None)

    def registrar_prestamo(self, data):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO prestamos (conductor, monto_total, fecha_prestamo, motivo, observaciones, estado, fecha_registro)
                VALUES (%s, %s, %s, %s, %s, 'activo', %s) RETURNING id
            """, (
                data['conductor'], int(data['monto_total']),
                data['fecha_prestamo'], data.get('motivo','').strip(),
                data.get('observaciones','').strip(), hora_colombia()
            ))
            nuevo_id = cur.fetchone()[0]
            conn.commit()
            return nuevo_id
        except Exception as e:
            conn.rollback()
            st.error(f"Error registrando préstamo: {e}")
            return None
        finally:
            put_conn(conn)

    def eliminar_prestamo(self, prestamo_id):
        self._exec("DELETE FROM prestamos WHERE id = %s", (prestamo_id,))

    def actualizar_estado_prestamo(self, prestamo_id, estado):
        self._exec("UPDATE prestamos SET estado = %s WHERE id = %s", (estado, prestamo_id))

    # ---- Pagos préstamos ----
    def obtener_pagos(self, prestamo_id=None):
        if prestamo_id:
            return self._query_df(
                "SELECT * FROM pagos_prestamos WHERE prestamo_id = %s ORDER BY fecha_pago DESC",
                (prestamo_id,)
            )
        return self._query_df("SELECT * FROM pagos_prestamos ORDER BY fecha_pago DESC")

    def registrar_pago(self, data):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO pagos_prestamos (prestamo_id, monto_pago, fecha_pago,
                    observaciones, registrado_por, fecha_registro)
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
            """, (
                data['prestamo_id'], int(data['monto_pago']),
                data['fecha_pago'], data.get('observaciones','').strip(),
                data.get('registrado_por','').strip().upper(), hora_colombia()
            ))
            nuevo_id = cur.fetchone()[0]
            conn.commit()
            return nuevo_id
        except Exception as e:
            conn.rollback()
            st.error(f"Error registrando pago: {e}")
            return None
        finally:
            put_conn(conn)

    def eliminar_pago(self, pago_id):
        self._exec("DELETE FROM pagos_prestamos WHERE id = %s", (pago_id,))

    # ---- Anticipos ----
    def registrar_viaje(self, data):
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO anticipos_v1
                (fecha_viaje, fecha_registro, placa, conductor, cliente,
                 origen, destino, valor_anticipo, observaciones, manifiesto, legalizado)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE) RETURNING id
            """, (
                data['fecha_viaje'], hora_colombia(), data['placa'], data['conductor'],
                data['cliente'], data['origen'], data['destino'],
                int(data['valor_anticipo']), data.get('observaciones',''),
                data.get('manifiesto','').strip().upper()
            ))
            nuevo_id = cur.fetchone()[0]
            conn.commit()
            return nuevo_id
        except Exception as e:
            conn.rollback()
            st.error(f"Error guardando viaje: {e}")
            return None
        finally:
            put_conn(conn)

    def editar_viaje(self, viaje_id, data):
        return bool(self._exec("""
            UPDATE anticipos_v1 SET
                fecha_viaje = %s, placa = %s, conductor = %s, cliente = %s,
                origen = %s, destino = %s, valor_anticipo = %s,
                observaciones = %s, manifiesto = %s
            WHERE id = %s
        """, (
            data['fecha_viaje'], data['placa'], data['conductor'], data['cliente'],
            data['origen'], data['destino'], int(data['valor_anticipo']),
            data.get('observaciones',''), data.get('manifiesto','').strip().upper(), viaje_id
        )))

    def legalizar(self, viaje_id, nombre_quien_legaliza, obs_legalizacion=""):
        return bool(self._exec("""
            UPDATE anticipos_v1
            SET legalizado = TRUE, fecha_legalizacion = %s,
                legalizado_por = %s, obs_legalizacion = %s
            WHERE id = %s
        """, (hora_colombia(), nombre_quien_legaliza, obs_legalizacion, viaje_id)))

    def buscar(self, estado=None, fecha_ini=None, fecha_fin=None, placa=None, conductor=None, manifiesto=None):
        q = "SELECT * FROM anticipos_v1 WHERE 1=1"
        params = []
        if estado == "legalizado":   q += " AND legalizado = TRUE"
        elif estado == "pendiente":  q += " AND legalizado = FALSE"
        if fecha_ini: q += " AND fecha_viaje >= %s"; params.append(fecha_ini)
        if fecha_fin: q += " AND fecha_viaje <= %s"; params.append(fecha_fin)
        if placa:     q += " AND placa = %s";        params.append(placa)
        if conductor: q += " AND conductor ILIKE %s"; params.append(f"%{conductor}%")
        if manifiesto:q += " AND manifiesto ILIKE %s"; params.append(f"%{manifiesto}%")
        q += " ORDER BY fecha_viaje DESC, fecha_registro DESC"
        return self._query_df(q, params if params else None)

    def eliminar(self, viaje_id):
        self._exec("DELETE FROM anticipos_v1 WHERE id = %s", (viaje_id,))

    def obtener_por_id(self, viaje_id):
        df = self._query_df("SELECT * FROM anticipos_v1 WHERE id = %s", (viaje_id,))
        return df.iloc[0] if not df.empty else None


# ==================== HELPERS ====================
def get_lista_clientes(db):
    extras_df = db.obtener_clientes_extra()
    extras = extras_df['nombre'].tolist() if not extras_df.empty else []
    return sorted(set(CLIENTES_DEFAULT + extras))

def get_lista_conductores(db):
    extras_df = db.obtener_conductores_extra()
    extras = extras_df['nombre'].tolist() if not extras_df.empty else []
    return sorted(set(CONDUCTORES_DEFAULT + extras))

def calcular_saldo_prestamo(prestamo_id, monto_total, df_pagos):
    if df_pagos.empty:
        return 0, int(monto_total)
    pagos_p = df_pagos[df_pagos["prestamo_id"] == prestamo_id]
    pagado = int(pagos_p["monto_pago"].sum())
    saldo  = max(0, int(monto_total) - pagado)
    return pagado, saldo


# ==================== APP PRINCIPAL ====================
def main():
    st.set_page_config(page_title="Anticipos - Transporte de Carga", layout="wide", page_icon="🚛")
    st.title("🚛 Gestión de Anticipos - Transporte de Carga")

    session_defaults = {
        'db': None, 'confirmar_eliminar': None, 'editando_id': None,
        'confirmar_eliminar_cliente': None, 'confirmar_eliminar_conductor': None,
        'editando_conductor_id': None, 'confirmar_eliminar_vac': None,
        'editando_vac_id': None,
        'confirmar_eliminar_prestamo': None, 'confirmar_eliminar_pago': None,
        'pago_vac_periodo': None,
        'confirmar_eliminar_pago_vac': None,
    }
    for key, val in session_defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val

    if st.session_state.db is None:
        db = DB()
        db.init_tablas()
        st.session_state.db = db
    db = st.session_state.db

    (tab_reg, tab_leg, tab_hist,
     tab_vac, tab_prest,
     tab_clientes, tab_conductores) = st.tabs([
        "📝 Registrar Viaje",
        "✅ Legalizar Anticipos",
        "📋 Historial",
        "🏖️ Vacaciones",
        "💰 Préstamos",
        "🏢 Clientes",
        "👤 Conductores",
    ])

    # ==================== TAB 1: REGISTRAR ====================
    with tab_reg:
        st.header("Registrar nuevo viaje con anticipo")
        lista_clientes    = get_lista_clientes(db)
        lista_conductores = get_lista_conductores(db)

        with st.form("form_registro", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                fecha_viaje = st.date_input("Fecha del viaje", value=datetime.today())
                placa       = st.selectbox("Placa de la tractomula", PLACAS)
                conductor   = st.selectbox("Conductor", lista_conductores)
                cliente     = st.selectbox("Cliente", lista_clientes)
            with col2:
                manifiesto  = st.text_input("Número de manifiesto ✱", placeholder="Ej: 1234567")
                origen      = st.text_input("Origen",  placeholder="Ciudad de origen")
                destino     = st.text_input("Destino", placeholder="Ciudad de destino")
                anticipo_txt = st.text_input("Valor del anticipo (COP)", placeholder="Ejemplo: 1.500.000")
                anticipo = limpiar(anticipo_txt)
                if anticipo > 0:
                    st.caption(f"💵 {fmt(anticipo)} COP")
                observaciones = st.text_area("Observaciones", height=80)

            submitted = st.form_submit_button("💾 Registrar Viaje", type="primary")
            if submitted:
                errores = []
                if not manifiesto.strip(): errores.append("El número de manifiesto es obligatorio")
                if not origen.strip():     errores.append("Origen es obligatorio")
                if not destino.strip():    errores.append("Destino es obligatorio")
                if anticipo <= 0:          errores.append("El valor del anticipo debe ser mayor a 0")
                if errores:
                    for e in errores: st.error(f"⚠️ {e}")
                else:
                    nuevo_id = db.registrar_viaje({
                        'fecha_viaje': fecha_viaje, 'placa': placa,
                        'conductor': conductor.strip().upper(),
                        'cliente': cliente.strip().upper(),
                        'origen': origen.strip().upper(),
                        'destino': destino.strip().upper(),
                        'valor_anticipo': anticipo,
                        'observaciones': observaciones.strip(),
                        'manifiesto': manifiesto.strip()
                    })
                    if nuevo_id:
                        st.success(f"✅ Viaje registrado exitosamente (ID: {nuevo_id})")

    # ==================== TAB 2: LEGALIZAR ====================
    with tab_leg:
        st.header("Legalizar anticipos pendientes")
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        with col_f1: fecha_ini_leg = st.date_input("Desde", value=None, key="leg_fi")
        with col_f2: fecha_fin_leg = st.date_input("Hasta", value=None, key="leg_ff")
        with col_f3: placa_leg = st.selectbox("Placa", ["Todas"] + PLACAS, key="leg_placa")
        with col_f4: manifiesto_leg = st.text_input("Buscar por manifiesto", placeholder="Nº manifiesto...", key="leg_manif")

        fi = fecha_ini_leg.strftime('%Y-%m-%d') if fecha_ini_leg else None
        ff = fecha_fin_leg.strftime('%Y-%m-%d') if fecha_fin_leg else None
        pl = None if placa_leg == "Todas" else placa_leg
        mf = manifiesto_leg.strip() if manifiesto_leg else None
        df_pendientes = db.buscar(estado="pendiente", fecha_ini=fi, fecha_fin=ff, placa=pl, manifiesto=mf)

        if df_pendientes.empty:
            st.success("✅ No hay anticipos pendientes de legalización.")
        else:
            criticos, atencion, al_dia = [], [], []
            for _, row in df_pendientes.iterrows():
                dias, nivel = clasificar_alerta(row['fecha_viaje'])
                entry = (row['id'], dias)
                if nivel == "critical": criticos.append(entry)
                elif nivel == "warning": atencion.append(entry)
                else: al_dia.append(entry)
            total_pendiente = df_pendientes['valor_anticipo'].sum()
            if criticos:
                st.error(f"🚨 **{len(criticos)} anticipo(s) CRÍTICO(S)** | 🟡 {len(atencion)} en atención | 🟢 {len(al_dia)} al día | 💰 Total: **${fmt(total_pendiente)} COP**")
            elif atencion:
                st.warning(f"🟡 **{len(atencion)} anticipo(s)** requieren atención | 🟢 {len(al_dia)} al día | 💰 Total: **${fmt(total_pendiente)} COP**")
            else:
                st.info(f"🟢 {len(al_dia)} viaje(s) pendiente(s), todos al día | 💰 Total: **${fmt(total_pendiente)} COP**")

            df_ordenado = df_pendientes.sort_values("fecha_viaje", ascending=False)
            for _, row in df_ordenado.iterrows():
                dias, nivel = clasificar_alerta(row['fecha_viaje'])
                badge = badge_alerta(dias, nivel)
                manif_label = f"Manif: {row.get('manifiesto','—')} | " if row.get('manifiesto') else ""
                label_expander = (f"{badge} | ID {row['id']} | {manif_label}"
                    f"{fmt_fecha(row['fecha_viaje'])} | {row['placa']} | {row['conductor']} | "
                    f"{row['origen']} → {row['destino']} | ${fmt(row['valor_anticipo'])} COP")
                with st.expander(label_expander):
                    col_info, col_form = st.columns([2, 2])
                    with col_info:
                        st.markdown("**Datos del viaje:**")
                        if nivel == "critical": st.error(f"⏰ Este anticipo lleva **{dias} días** sin legalizar")
                        elif nivel == "warning": st.warning(f"⚠️ Este anticipo lleva **{dias} días** sin legalizar")
                        else: st.success(f"✅ {dias} días desde el viaje — al día")
                        st.write(f"📄 Manifiesto: **{row.get('manifiesto', '—')}**")
                        st.write(f"📅 Fecha: {fmt_fecha(row['fecha_viaje'])} | 🚛 {row['placa']} | 👤 {row['conductor']}")
                        st.write(f"🏢 {row['cliente']} | 📍 {row['origen']} → {row['destino']}")
                        st.write(f"💰 **${fmt(row['valor_anticipo'])} COP**")
                    with col_form:
                        st.markdown("**Legalizar este viaje:**")
                        nombre_leg = st.text_input("Tu nombre completo", key=f"nombre_leg_{row['id']}")
                        obs_leg = st.text_area("Observaciones", height=80, key=f"obs_leg_{row['id']}")
                        if st.button("✅ Marcar como LEGALIZADO", key=f"btn_leg_{row['id']}", type="primary"):
                            if not nombre_leg.strip():
                                st.error("⚠️ Debes escribir tu nombre.")
                            else:
                                ok = db.legalizar(row['id'], nombre_leg.strip().upper(), obs_leg.strip())
                                if ok:
                                    st.success(f"✅ Viaje ID {row['id']} legalizado."); st.rerun()

    # ==================== TAB 3: HISTORIAL ====================
    with tab_hist:
        st.header("Historial de viajes")
        col1, col2, col3 = st.columns(3)
        with col1: estado_filtro = st.selectbox("Estado", ["Todos","Pendientes","Legalizados"], key="hist_estado")
        with col2: fecha_ini_h  = st.date_input("Desde", value=None, key="hist_fi")
        with col3: fecha_fin_h  = st.date_input("Hasta", value=None, key="hist_ff")
        col4, col5, col6 = st.columns(3)
        with col4: placa_h      = st.selectbox("Placa", ["Todas"] + PLACAS, key="hist_placa")
        with col5: conductor_h  = st.text_input("Buscar conductor", key="hist_cond")
        with col6: manifiesto_h = st.text_input("Buscar por manifiesto", key="hist_manif")

        estado_map = {"Todos": None, "Pendientes": "pendiente", "Legalizados": "legalizado"}
        fi_h  = fecha_ini_h.strftime('%Y-%m-%d') if fecha_ini_h else None
        ff_h  = fecha_fin_h.strftime('%Y-%m-%d') if fecha_fin_h else None
        pl_h  = None if placa_h == "Todas" else placa_h
        df_hist = db.buscar(estado=estado_map[estado_filtro], fecha_ini=fi_h, fecha_fin=ff_h,
                            placa=pl_h, conductor=conductor_h or None, manifiesto=manifiesto_h.strip() or None)

        if df_hist.empty:
            st.info("No se encontraron viajes con los filtros aplicados.")
        else:
            total_anticipo = df_hist['valor_anticipo'].sum()
            legalizados = int(df_hist['legalizado'].sum())
            pendientes  = len(df_hist) - legalizados
            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            col_m1.metric("Total viajes",    len(df_hist))
            col_m2.metric("Legalizados",     legalizados)
            col_m3.metric("Pendientes",      pendientes)
            col_m4.metric("Total anticipos", f"${fmt(total_anticipo)}")

            cols_tabla = ['id','manifiesto','fecha_viaje','placa','conductor','cliente','origen',
                          'destino','valor_anticipo','legalizado','legalizado_por','fecha_legalizacion']
            df_show = df_hist[[c for c in cols_tabla if c in df_hist.columns]].copy()
            df_show['dias_alerta'] = df_hist.apply(
                lambda r: "—" if r.get('legalizado') else badge_alerta(*clasificar_alerta(r['fecha_viaje'])), axis=1)
            df_show['valor_anticipo'] = df_show['valor_anticipo'].apply(lambda x: f"${fmt(x)}")
            df_show['legalizado'] = df_show['legalizado'].apply(lambda x: "✅ Legalizado" if x else "🔴 Pendiente")
            # Formatear fechas a DD/MM/YYYY
            df_show['fecha_viaje'] = df_show['fecha_viaje'].apply(fmt_fecha)
            df_show['fecha_legalizacion'] = df_show['fecha_legalizacion'].apply(
                lambda x: fmt_fecha(x) if pd.notna(x) and str(x) not in ['', 'None', 'NaT'] else "—"
            )
            df_show.rename(columns={
                'id':'ID','manifiesto':'Manifiesto','fecha_viaje':'Fecha viaje','placa':'Placa',
                'conductor':'Conductor','cliente':'Cliente','origen':'Origen','destino':'Destino',
                'valor_anticipo':'Anticipo','legalizado':'Estado','legalizado_por':'Legalizado por',
                'fecha_legalizacion':'Fecha legalización','dias_alerta':'Alerta'
            }, inplace=True)
            st.dataframe(df_show, use_container_width=True, hide_index=True, height=350)

            st.divider()
            col_exp1, col_exp2 = st.columns([3, 1])
            with col_exp1:
                titulo_excel = st.text_input("Título del reporte Excel",
                    value=f"Anticipos {estado_filtro} — {hora_colombia().strftime('%d/%m/%Y')}", key="titulo_excel")
            with col_exp2:
                st.markdown("&nbsp;")
                excel_bytes = generar_excel(df_hist, titulo_excel)
                st.download_button(label="📥 Exportar a Excel", data=excel_bytes,
                    file_name=f"anticipos_{hora_colombia().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

            st.divider()
            st.subheader("Acciones sobre un viaje")
            viaje_sel = st.selectbox("Selecciona un viaje por ID", df_hist['id'].tolist(),
                format_func=lambda x: (
                    f"ID {x} | Manif: {df_hist[df_hist['id']==x]['manifiesto'].values[0] or '—'} | "
                    f"{df_hist[df_hist['id']==x]['placa'].values[0]} | {df_hist[df_hist['id']==x]['conductor'].values[0]}"
                ), key="hist_sel")
            row_sel = df_hist[df_hist['id'] == viaje_sel].iloc[0]
            col_det, col_acc = st.columns([3, 1])
            with col_det:
                estado_tag = "✅ **LEGALIZADO**" if row_sel['legalizado'] else "🔴 **PENDIENTE**"
                st.markdown(f"**Estado:** {estado_tag}")
                st.write(f"📄 Manifiesto: **{row_sel.get('manifiesto', '—')}**")
                st.write(f"Fecha: {fmt_fecha(row_sel['fecha_viaje'])} | Placa: {row_sel['placa']} | Conductor: {row_sel['conductor']}")
                st.write(f"Ruta: {row_sel['origen']} → {row_sel['destino']} | Anticipo: **${fmt(row_sel['valor_anticipo'])} COP**")
                if row_sel['legalizado']:
                    st.success(f"Legalizado por: **{row_sel['legalizado_por']}** | Fecha: {fmt_fecha(row_sel['fecha_legalizacion'])}")
            with col_acc:
                if st.button("✏️ Editar viaje", key="btn_editar"):
                    st.session_state.editando_id = viaje_sel; st.rerun()
                if st.session_state.confirmar_eliminar == viaje_sel:
                    st.warning(f"¿Eliminar ID **{viaje_sel}**?")
                    c_si, c_no = st.columns(2)
                    with c_si:
                        if st.button("Sí", key="btn_si_eliminar", type="primary"):
                            db.eliminar(viaje_sel); st.session_state.confirmar_eliminar = None
                            st.success("Eliminado."); st.rerun()
                    with c_no:
                        if st.button("Cancelar", key="btn_no_eliminar"):
                            st.session_state.confirmar_eliminar = None; st.rerun()
                else:
                    if st.button("🗑️ Eliminar viaje", key="btn_eliminar", type="secondary"):
                        st.session_state.confirmar_eliminar = viaje_sel; st.rerun()

        if st.session_state.editando_id is not None:
            eid = st.session_state.editando_id
            viaje_edit = db.obtener_por_id(eid)
            if viaje_edit is not None:
                st.divider()
                st.subheader(f"✏️ Editando viaje ID {eid}")
                lista_clientes_edit = get_lista_clientes(db)
                cliente_actual = viaje_edit['cliente']
                if cliente_actual not in lista_clientes_edit: lista_clientes_edit = [cliente_actual] + lista_clientes_edit
                idx_cliente = lista_clientes_edit.index(cliente_actual)
                lista_conductores_edit = get_lista_conductores(db)
                conductor_actual = viaje_edit['conductor']
                if conductor_actual not in lista_conductores_edit: lista_conductores_edit = [conductor_actual] + lista_conductores_edit
                idx_conductor_edit = lista_conductores_edit.index(conductor_actual)
                idx_placa_edit = PLACAS.index(viaje_edit['placa']) if viaje_edit['placa'] in PLACAS else 0

                with st.form(f"form_editar_{eid}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        fecha_e     = st.date_input("Fecha del viaje", value=pd.to_datetime(viaje_edit['fecha_viaje']).date())
                        placa_e     = st.selectbox("Placa", PLACAS, index=idx_placa_edit)
                        conductor_e = st.selectbox("Conductor", lista_conductores_edit, index=idx_conductor_edit)
                        cliente_e   = st.selectbox("Cliente", lista_clientes_edit, index=idx_cliente)
                        manifiesto_e = st.text_input("Número de manifiesto ✱", value=viaje_edit.get('manifiesto','') or '')
                    with col2:
                        origen_e  = st.text_input("Origen",  value=viaje_edit['origen'])
                        destino_e = st.text_input("Destino", value=viaje_edit['destino'])
                        anticipo_e_txt = st.text_input("Valor del anticipo (COP)", value=fmt(viaje_edit['valor_anticipo']))
                        anticipo_e = limpiar(anticipo_e_txt)
                        obs_e = st.text_area("Observaciones", value=viaje_edit.get('observaciones','') or '', height=80)
                    col_g, col_c = st.columns(2)
                    with col_g: guardar_edit  = st.form_submit_button("💾 Guardar cambios", type="primary")
                    with col_c: cancelar_edit = st.form_submit_button("✖ Cancelar")
                    if guardar_edit:
                        errores_e = []
                        if not manifiesto_e.strip(): errores_e.append("Manifiesto obligatorio")
                        if not origen_e.strip():     errores_e.append("Origen obligatorio")
                        if not destino_e.strip():    errores_e.append("Destino obligatorio")
                        if anticipo_e <= 0:          errores_e.append("Anticipo debe ser mayor a 0")
                        if errores_e:
                            for err in errores_e: st.error(f"⚠️ {err}")
                        else:
                            ok = db.editar_viaje(eid, {
                                'fecha_viaje': fecha_e, 'placa': placa_e,
                                'conductor': conductor_e.strip().upper(),
                                'cliente': cliente_e.strip().upper(),
                                'origen': origen_e.strip().upper(),
                                'destino': destino_e.strip().upper(),
                                'valor_anticipo': anticipo_e,
                                'observaciones': obs_e.strip(),
                                'manifiesto': manifiesto_e.strip()
                            })
                            if ok:
                                st.success(f"✅ Viaje ID {eid} actualizado."); st.session_state.editando_id = None; st.rerun()
                    if cancelar_edit:
                        st.session_state.editando_id = None; st.rerun()

    # ==================== TAB 4: VACACIONES v15 ====================
    with tab_vac:
        st.header("🏖️ Control de Vacaciones")
        st.caption(f"Regla: {DIAS_VACACIONES_ANUALES} días de vacaciones por cada año laboral completado.")

        lista_conductores_vac = get_lista_conductores(db)
        df_info_todos         = db.obtener_todos_info_conductores()
        df_vac_todos          = db.obtener_vacaciones()
        df_pagos_vac_todos    = db.obtener_pagos_vacaciones()
        hoy                   = hora_colombia().date()

        v_tab1, v_tab2, v_tab3 = st.tabs([
            "📊 Resumen y saldos",
            "📝 Registrar vacaciones tomadas",
            "⚙️ Fecha de ingreso"
        ])

        # ======================== RESUMEN Y SALDOS ========================
        with v_tab1:

            col_f1v, col_f2v = st.columns(2)
            with col_f1v:
                filtro_cond_vac = st.selectbox(
                    "Filtrar por conductor",
                    ["Todos"] + lista_conductores_vac,
                    key="vac_filtro_cond"
                )
            with col_f2v:
                filtro_estado_vac = st.selectbox(
                    "Filtrar por estado",
                    ["Todos", "🔴 Con días pendientes", "✅ Al día / Sin períodos", "⚪ Sin fecha ingreso"],
                    key="vac_filtro_estado"
                )

            conductores_mostrar = lista_conductores_vac if filtro_cond_vac == "Todos" else [filtro_cond_vac]

            filas = []
            for cond in conductores_mostrar:
                info_row = df_info_todos[df_info_todos["conductor"] == cond].iloc[0] \
                    if not df_info_todos.empty and (df_info_todos["conductor"] == cond).any() else None

                if info_row is not None and info_row.get("fecha_ingreso") is not None:
                    fi = pd.to_datetime(info_row["fecha_ingreso"]).date()
                    calc = calcular_vacaciones(cond, fi, df_vac_todos, hoy)

                    pagos_cond_df = df_pagos_vac_todos[df_pagos_vac_todos["conductor"] == cond] \
                        if not df_pagos_vac_todos.empty else pd.DataFrame()
                    anios_pagados_dinero = set(pagos_cond_df["anio_laboral"].tolist()) \
                        if not pagos_cond_df.empty else set()

                    hay_pendiente_real = any(
                        p["anio"] not in anios_pagados_dinero and (
                            DIAS_VACACIONES_ANUALES - sum(
                                int(r.get("dias", 0)) for r in calc["registros"]
                                if r.get("fecha_inicio") is not None and (
                                    pd.to_datetime(r["fecha_inicio"]).date() >= p["inicio"]
                                    and pd.to_datetime(r["fecha_inicio"]).date() < p["fin"]
                                )
                            )
                        ) > 0
                        for p in calc["periodos"]
                    )

                    if hay_pendiente_real:
                        estado_v = "pendientes"
                    elif calc["anios_completos"] > 0:
                        estado_v = "al_dia"
                    else:
                        estado_v = "sin_periodos"

                    filas.append({
                        "conductor": cond,
                        "fecha_ingreso": fi,
                        "calc": calc,
                        "estado_v": estado_v,
                        "anios_pagados_dinero": anios_pagados_dinero,
                        "pagos_cond_df": pagos_cond_df,
                    })
                else:
                    filas.append({
                        "conductor": cond,
                        "fecha_ingreso": None,
                        "calc": None,
                        "estado_v": "sin_fecha",
                        "anios_pagados_dinero": set(),
                        "pagos_cond_df": pd.DataFrame(),
                    })

            if filtro_estado_vac == "🔴 Con días pendientes":
                filas = [f for f in filas if f["estado_v"] == "pendientes"]
            elif filtro_estado_vac == "✅ Al día / Sin períodos":
                filas = [f for f in filas if f["estado_v"] in ("al_dia", "sin_periodos")]
            elif filtro_estado_vac == "⚪ Sin fecha ingreso":
                filas = [f for f in filas if f["estado_v"] == "sin_fecha"]

            con_pendientes = sum(1 for f in filas if f["estado_v"] == "pendientes")
            al_dia_count   = sum(1 for f in filas if f["estado_v"] in ("al_dia", "sin_periodos"))
            sin_fecha_c    = sum(1 for f in filas if f["estado_v"] == "sin_fecha")
            total_dias_pendientes = sum(
                f["calc"]["dias_vencidos"] for f in filas
                if f["calc"] and f["calc"]["dias_vencidos"] > 0
            )

            col_mv1, col_mv2, col_mv3, col_mv4 = st.columns(4)
            col_mv1.metric("Conductores mostrados", len(filas))
            col_mv2.metric("🔴 Con días pendientes", con_pendientes)
            col_mv3.metric("✅ Al día",              al_dia_count)
            col_mv4.metric("Días totales pendientes", total_dias_pendientes)

            if con_pendientes > 0:
                nombres_pendientes = [f["conductor"] for f in filas if f["estado_v"] == "pendientes"]
                st.error(f"🚨 {con_pendientes} conductor(es) con vacaciones pendientes: " + ", ".join(nombres_pendientes))

            st.divider()

            for fila in filas:
                cond              = fila["conductor"]
                calc              = fila["calc"]
                estado_v          = fila["estado_v"]
                anios_pagados_din = fila["anios_pagados_dinero"]
                pagos_cond_df     = fila["pagos_cond_df"]

                if estado_v == "pendientes":
                    label_v = (
                        f"🔴  {cond}  |  "
                        f"{calc['dias_vencidos']} días pendientes  |  "
                        f"{calc['dias_generados']} generados — {calc['dias_usados']} usados  |  "
                        f"Próxima vacación: {calc['proxima_fecha'].strftime('%d/%m/%Y')}"
                    )
                elif estado_v == "al_dia":
                    label_v = (
                        f"✅  {cond}  |  Al día  |  "
                        f"{calc['dias_generados']} generados — {calc['dias_usados']} usados  |  "
                        f"Próxima vacación: {calc['proxima_fecha'].strftime('%d/%m/%Y')} "
                        f"(en {calc['dias_para_proxima']} días)"
                    )
                elif estado_v == "sin_periodos":
                    label_v = (
                        f"🟢  {cond}  |  Menos de 1 año  |  "
                        f"Primera vacación: {calc['proxima_fecha'].strftime('%d/%m/%Y')} "
                        f"(en {calc['dias_para_proxima']} días)"
                    )
                else:
                    label_v = f"⚪  {cond}  |  Sin fecha de ingreso registrada"

                with st.expander(label_v):

                    if estado_v == "sin_fecha":
                        st.warning("⚠️ Registra la fecha de ingreso del conductor en la pestaña **⚙️ Fecha de ingreso**.")
                        continue

                    col_va, col_vb = st.columns([3, 2])

                    with col_va:
                        st.write(f"**Ingreso:** {fila['fecha_ingreso'].strftime('%d/%m/%Y')}")

                        m1, m2, m3, m4 = st.columns(4)
                        m1.metric("Años completados",  calc["anios_completos"])
                        m2.metric("Días generados",    calc["dias_generados"])
                        m3.metric("Días usados",       calc["dias_usados"])
                        if calc["dias_vencidos"] > 0:
                            m4.metric("🔴 Días pendientes", calc["dias_vencidos"])
                        else:
                            m4.metric("✅ Días pendientes", 0)

                        if estado_v == "pendientes":
                            st.error(
                                f"⏰ Este conductor tiene **{calc['dias_vencidos']} días** de vacaciones "
                                f"sin tomar, generados en {calc['anios_completos']} año(s) laborales."
                            )
                        elif calc["anios_completos"] > 0:
                            st.success("✅ Todos los días de vacaciones están utilizados o pagados.")
                        else:
                            st.info("⏳ Aún no completa su primer año laboral.")

                        if calc["dias_para_proxima"] <= 30:
                            st.warning(
                                f"📆 Próxima vacación: **{calc['proxima_fecha'].strftime('%d/%m/%Y')}** "
                                f"— ¡faltan solo **{calc['dias_para_proxima']} días**!"
                            )
                        else:
                            st.info(
                                f"📆 Próxima vacación: **{calc['proxima_fecha'].strftime('%d/%m/%Y')}** "
                                f"(en {calc['dias_para_proxima']} días)"
                            )

                        # ── Cronología de períodos — lógica v15 ──
                        # REGLAS:
                        # ✅ verde  → días tomados == 15 Y pago en dinero registrado (o días == 15 sin necesidad de pago)
                        # 🟡 amarillo → días tomados > 0 pero < 15, sin pago registrado → "Vacaciones parciales, falta pago"
                        # 💵 azul    → pago en dinero registrado, pero aún faltan vacaciones físicas por tomar
                        # 🔴 rojo    → 0 días tomados y sin pago → completamente pendiente
                        if calc["periodos"]:
                            st.markdown("**Períodos anuales completados:**")
                            for p in calc["periodos"]:
                                anio_num = p["anio"]

                                regs_periodo = [
                                    r for r in calc["registros"]
                                    if r.get("fecha_inicio") is not None and (
                                        pd.to_datetime(r["fecha_inicio"]).date() >= p["inicio"]
                                        and pd.to_datetime(r["fecha_inicio"]).date() < p["fin"]
                                    )
                                ]
                                dias_tomados  = sum(int(r.get("dias", 0)) for r in regs_periodo)
                                dias_pend_p   = DIAS_VACACIONES_ANUALES - dias_tomados

                                pago_row = pagos_cond_df[pagos_cond_df["anio_laboral"] == anio_num] \
                                    if not pagos_cond_df.empty else pd.DataFrame()
                                pagado_en_dinero = not pago_row.empty
                                monto_pago_din   = int(pago_row.iloc[0]["monto_cop"]) if pagado_en_dinero else 0
                                fecha_pago_din   = fmt_fecha(pago_row.iloc[0]["fecha_pago"]) if pagado_en_dinero else ""
                                reg_por_din      = pago_row.iloc[0].get("registrado_por","") if pagado_en_dinero else ""
                                pago_vac_id      = int(pago_row.iloc[0]["id"]) if pagado_en_dinero else None

                                # ── CASO 1: Todos los días tomados físicamente Y sin días restantes ──
                                if dias_pend_p <= 0 and not pagado_en_dinero:
                                    st.success(
                                        f"✅ **{p['label']}** — "
                                        f"{dias_tomados} días tomados físicamente. Período completo."
                                    )

                                # ── CASO 2: Todos los días tomados físicamente Y hay pago (redundante pero posible) ──
                                elif dias_pend_p <= 0 and pagado_en_dinero:
                                    st.success(
                                        f"✅ **{p['label']}** — "
                                        f"{dias_tomados} días tomados. Pago registrado: ${fmt(monto_pago_din)} COP "
                                        f"({fecha_pago_din}). Período completo."
                                    )

                                # ── CASO 3: Hay pago en dinero registrado, pero AÚN faltan días físicos ──
                                elif pagado_en_dinero and dias_pend_p > 0:
                                    st.info(
                                        f"💵 **{p['label']}** — "
                                        f"{dias_tomados} días tomados físicamente · "
                                        f"**{dias_pend_p} días pendientes de tomar** · "
                                        f"Pago en dinero: ${fmt(monto_pago_din)} COP · "
                                        f"Fecha pago: {fecha_pago_din}"
                                        + (f" · Registrado por: {reg_por_din}" if reg_por_din else "")
                                    )
                                    st.caption(
                                        "ℹ️ Se registró pago en dinero, pero aún faltan vacaciones físicas por tomar. "
                                        "Puedes eliminar el pago si fue un error."
                                    )
                                    if st.session_state.confirmar_eliminar_pago_vac == pago_vac_id:
                                        st.warning(f"¿Eliminar el pago de **{p['label']}**? Quedará como pendiente.")
                                        col_si_pv, col_no_pv, _ = st.columns([1, 1, 4])
                                        with col_si_pv:
                                            if st.button("✅ Sí", key=f"si_pvac_{pago_vac_id}", type="primary"):
                                                db.eliminar_pago_vacacion(pago_vac_id)
                                                st.session_state.confirmar_eliminar_pago_vac = None
                                                st.rerun()
                                        with col_no_pv:
                                            if st.button("❌ No", key=f"no_pvac_{pago_vac_id}"):
                                                st.session_state.confirmar_eliminar_pago_vac = None
                                                st.rerun()
                                    else:
                                        if st.button(
                                            "🗑️ Eliminar pago",
                                            key=f"del_pvac_{cond}_{anio_num}",
                                            help="Eliminar pago en dinero (revertir a pendiente)"
                                        ):
                                            st.session_state.confirmar_eliminar_pago_vac = pago_vac_id
                                            st.rerun()

                                # ── CASO 4: Días parcialmente tomados, SIN pago registrado ──
                                elif dias_tomados > 0 and dias_pend_p > 0 and not pagado_en_dinero:
                                    st.warning(
                                        f"🟡 **{p['label']}** — "
                                        f"{dias_tomados} días tomados · "
                                        f"**{dias_pend_p} días pendientes** · "
                                        f"Faltan vacaciones o pago en dinero"
                                    )
                                    clave_form = (cond, anio_num)
                                    if st.session_state.pago_vac_periodo != clave_form:
                                        if st.button(
                                            "💵 Registrar pago en dinero",
                                            key=f"btn_pago_vac_{cond}_{anio_num}",
                                            type="primary"
                                        ):
                                            st.session_state.pago_vac_periodo = clave_form
                                            st.rerun()

                                    if st.session_state.pago_vac_periodo == (cond, anio_num):
                                        with st.form(f"form_pago_vac_{cond}_{anio_num}"):
                                            st.markdown(
                                                f"**💵 Registrar pago en dinero — {p['label']}**  \n"
                                                f"Días pendientes a compensar: **{dias_pend_p}**"
                                            )
                                            col_pv1, col_pv2 = st.columns(2)
                                            with col_pv1:
                                                monto_pv_txt = st.text_input(
                                                    "Monto pagado (COP)",
                                                    placeholder="Ej: 750.000",
                                                    key=f"monto_pv_{cond}_{anio_num}"
                                                )
                                                monto_pv = limpiar(monto_pv_txt)
                                                if monto_pv > 0:
                                                    st.caption(f"💵 {fmt(monto_pv)} COP")
                                                fecha_pv = st.date_input(
                                                    "Fecha del pago",
                                                    value=datetime.today(),
                                                    key=f"fecha_pv_{cond}_{anio_num}"
                                                )
                                            with col_pv2:
                                                reg_por_pv = st.text_input(
                                                    "Registrado por",
                                                    placeholder="Tu nombre completo",
                                                    key=f"reg_pv_{cond}_{anio_num}"
                                                )
                                                obs_pv = st.text_area(
                                                    "Observaciones",
                                                    height=68,
                                                    key=f"obs_pv_{cond}_{anio_num}"
                                                )

                                            col_gp, col_cp = st.columns(2)
                                            with col_gp:
                                                guardar_pv = st.form_submit_button("💾 Guardar pago", type="primary")
                                            with col_cp:
                                                cancelar_pv = st.form_submit_button("✖ Cancelar")

                                            if guardar_pv:
                                                errores_pv = []
                                                if monto_pv <= 0:
                                                    errores_pv.append("El monto debe ser mayor a 0.")
                                                if not reg_por_pv.strip():
                                                    errores_pv.append("Ingresa tu nombre.")
                                                if errores_pv:
                                                    for ep in errores_pv:
                                                        st.error(f"⚠️ {ep}")
                                                else:
                                                    nid_pv = db.registrar_pago_vacacion({
                                                        'conductor': cond,
                                                        'anio_laboral': anio_num,
                                                        'periodo_label': p['label'],
                                                        'monto_cop': monto_pv,
                                                        'fecha_pago': fecha_pv,
                                                        'observaciones': obs_pv.strip(),
                                                        'registrado_por': reg_por_pv.strip(),
                                                    })
                                                    if nid_pv:
                                                        st.success(
                                                            f"✅ Pago registrado para **{cond}** — "
                                                            f"{p['label']} — ${fmt(monto_pv)} COP"
                                                        )
                                                        st.session_state.pago_vac_periodo = None
                                                        st.rerun()
                                            if cancelar_pv:
                                                st.session_state.pago_vac_periodo = None
                                                st.rerun()

                                # ── CASO 5: 0 días tomados, sin pago → completamente pendiente ──
                                else:
                                    st.error(
                                        f"🔴 **{p['label']}** — "
                                        f"0 días tomados · "
                                        f"**{dias_pend_p} días pendientes** · "
                                        f"Sin vacaciones ni pago registrado"
                                    )
                                    clave_form = (cond, anio_num)
                                    if st.session_state.pago_vac_periodo != clave_form:
                                        if st.button(
                                            "💵 Registrar pago en dinero",
                                            key=f"btn_pago_vac_{cond}_{anio_num}",
                                            type="primary"
                                        ):
                                            st.session_state.pago_vac_periodo = clave_form
                                            st.rerun()

                                    if st.session_state.pago_vac_periodo == (cond, anio_num):
                                        with st.form(f"form_pago_vac_{cond}_{anio_num}"):
                                            st.markdown(
                                                f"**💵 Registrar pago en dinero — {p['label']}**  \n"
                                                f"Días a compensar: **{dias_pend_p}**"
                                            )
                                            col_pv1, col_pv2 = st.columns(2)
                                            with col_pv1:
                                                monto_pv_txt = st.text_input(
                                                    "Monto pagado (COP)",
                                                    placeholder="Ej: 750.000",
                                                    key=f"monto_pv_{cond}_{anio_num}"
                                                )
                                                monto_pv = limpiar(monto_pv_txt)
                                                if monto_pv > 0:
                                                    st.caption(f"💵 {fmt(monto_pv)} COP")
                                                fecha_pv = st.date_input(
                                                    "Fecha del pago",
                                                    value=datetime.today(),
                                                    key=f"fecha_pv_{cond}_{anio_num}"
                                                )
                                            with col_pv2:
                                                reg_por_pv = st.text_input(
                                                    "Registrado por",
                                                    placeholder="Tu nombre completo",
                                                    key=f"reg_pv_{cond}_{anio_num}"
                                                )
                                                obs_pv = st.text_area(
                                                    "Observaciones",
                                                    height=68,
                                                    key=f"obs_pv_{cond}_{anio_num}"
                                                )

                                            col_gp, col_cp = st.columns(2)
                                            with col_gp:
                                                guardar_pv = st.form_submit_button("💾 Guardar pago", type="primary")
                                            with col_cp:
                                                cancelar_pv = st.form_submit_button("✖ Cancelar")

                                            if guardar_pv:
                                                errores_pv = []
                                                if monto_pv <= 0:
                                                    errores_pv.append("El monto debe ser mayor a 0.")
                                                if not reg_por_pv.strip():
                                                    errores_pv.append("Ingresa tu nombre.")
                                                if errores_pv:
                                                    for ep in errores_pv:
                                                        st.error(f"⚠️ {ep}")
                                                else:
                                                    nid_pv = db.registrar_pago_vacacion({
                                                        'conductor': cond,
                                                        'anio_laboral': anio_num,
                                                        'periodo_label': p['label'],
                                                        'monto_cop': monto_pv,
                                                        'fecha_pago': fecha_pv,
                                                        'observaciones': obs_pv.strip(),
                                                        'registrado_por': reg_por_pv.strip(),
                                                    })
                                                    if nid_pv:
                                                        st.success(
                                                            f"✅ Pago registrado para **{cond}** — "
                                                            f"{p['label']} — ${fmt(monto_pv)} COP"
                                                        )
                                                        st.session_state.pago_vac_periodo = None
                                                        st.rerun()
                                            if cancelar_pv:
                                                st.session_state.pago_vac_periodo = None
                                                st.rerun()

                    with col_vb:
                        st.markdown("**Vacaciones registradas (tomadas):**")
                        df_vac_cond = df_vac_todos[df_vac_todos["conductor"] == cond] \
                            if not df_vac_todos.empty else pd.DataFrame()

                        if df_vac_cond.empty:
                            st.info("No hay vacaciones registradas para este conductor.")
                        else:
                            for _, vrow in df_vac_cond.sort_values("fecha_inicio", ascending=False).iterrows():
                                vid = vrow['id']

                                if st.session_state.editando_vac_id == vid:
                                    with st.form(f"form_edit_vac_{vid}"):
                                        st.markdown(f"**Editando ID {vid}**")
                                        col_ev1, col_ev2 = st.columns(2)
                                        with col_ev1:
                                            fi_edit = st.date_input(
                                                "Fecha inicio",
                                                value=pd.to_datetime(vrow['fecha_inicio']).date(),
                                                key=f"fi_edit_{vid}"
                                            )
                                            ff_edit = st.date_input(
                                                "Fecha fin",
                                                value=pd.to_datetime(vrow['fecha_fin']).date(),
                                                key=f"ff_edit_{vid}"
                                            )
                                        with col_ev2:
                                            dias_edit = st.number_input(
                                                "Días", min_value=1, max_value=60,
                                                value=min(int(vrow['dias']), 60),
                                                key=f"d_edit_{vid}"
                                            )
                                            obs_edit = st.text_area(
                                                "Observaciones",
                                                value=vrow.get('observaciones','') or '',
                                                height=60, key=f"obs_edit_{vid}"
                                            )
                                        col_ge, col_ce = st.columns(2)
                                        with col_ge:
                                            guardar_vac = st.form_submit_button("💾 Guardar", type="primary")
                                        with col_ce:
                                            cancelar_vac = st.form_submit_button("✖ Cancelar")

                                        if guardar_vac:
                                            ok_ev = db.actualizar_vacacion(vid, {
                                                'fecha_inicio': fi_edit,
                                                'fecha_fin': ff_edit,
                                                'dias': dias_edit,
                                                'observaciones': obs_edit.strip()
                                            })
                                            if ok_ev:
                                                st.session_state.editando_vac_id = None
                                                st.success("✅ Actualizado.")
                                                st.rerun()
                                        if cancelar_vac:
                                            st.session_state.editando_vac_id = None
                                            st.rerun()

                                else:
                                    fi_str  = fmt_fecha(vrow['fecha_inicio'])
                                    ff_str  = fmt_fecha(vrow['fecha_fin'])
                                    obs_str = vrow.get('observaciones','') or '—'
                                    st.write(
                                        f"📆 {fi_str} → {ff_str}  |  "
                                        f"**{vrow['dias']} días**  |  {obs_str}"
                                    )

                                    col_ve, col_vd = st.columns([1, 1])
                                    with col_ve:
                                        if st.button("✏️ Editar", key=f"edit_vac_{vid}"):
                                            st.session_state.editando_vac_id = vid
                                            st.rerun()
                                    with col_vd:
                                        if st.session_state.confirmar_eliminar_vac == vid:
                                            st.warning("¿Confirmar eliminación?")
                                            col_si_v, col_no_v = st.columns(2)
                                            with col_si_v:
                                                if st.button("✅ Sí", key=f"si_vac_{vid}", type="primary"):
                                                    db.eliminar_vacacion(vid)
                                                    st.session_state.confirmar_eliminar_vac = None
                                                    st.rerun()
                                            with col_no_v:
                                                if st.button("❌ No", key=f"no_vac_{vid}"):
                                                    st.session_state.confirmar_eliminar_vac = None
                                                    st.rerun()
                                        else:
                                            if st.button("🗑️ Eliminar", key=f"del_vac_{vid}"):
                                                st.session_state.confirmar_eliminar_vac = vid
                                                st.rerun()

                                    st.divider()

            st.divider()
            excel_vac = generar_excel_vacaciones(df_info_todos, df_vac_todos, df_pagos_vac_todos, lista_conductores_vac)
            st.download_button(
                label="📥 Exportar vacaciones a Excel",
                data=excel_vac,
                file_name=f"vacaciones_{hora_colombia().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

        # ======================== REGISTRAR VACACIONES TOMADAS ========================
        with v_tab2:
            st.subheader("Registrar vacaciones tomadas")
            st.caption("Aquí registras los días que el conductor YA tomó de vacaciones.")

            col_sel, _ = st.columns([2, 2])
            with col_sel:
                cond_sel = st.selectbox("Conductor", lista_conductores_vac, key="vac_cond_reg")

            info_cond = df_info_todos[df_info_todos["conductor"] == cond_sel].iloc[0] \
                if not df_info_todos.empty and (df_info_todos["conductor"] == cond_sel).any() else None

            if info_cond is None or info_cond.get("fecha_ingreso") is None:
                st.warning(f"⚠️ **{cond_sel}** no tiene fecha de ingreso. Regístrala primero en **⚙️ Fecha de ingreso**.")
            else:
                fi_cond = pd.to_datetime(info_cond["fecha_ingreso"]).date()
                calc_preview = calcular_vacaciones(cond_sel, fi_cond, df_vac_todos, hoy)

                col_p1, col_p2, col_p3, col_p4 = st.columns(4)
                col_p1.metric("Años completados",  calc_preview["anios_completos"])
                col_p2.metric("Días generados",    calc_preview["dias_generados"])
                col_p3.metric("Días usados",       calc_preview["dias_usados"])
                col_p4.metric("Días disponibles",  calc_preview["dias_vencidos"])

                if calc_preview["anios_completos"] == 0:
                    st.info(
                        f"⏳ **{cond_sel}** aún no ha completado su primer año. "
                        f"Primera vacación disponible: **{calc_preview['proxima_fecha'].strftime('%d/%m/%Y')}**"
                    )
                elif calc_preview["dias_vencidos"] == 0:
                    st.success(f"✅ **{cond_sel}** tiene todos sus días de vacaciones utilizados.")
                else:
                    st.warning(
                        f"⚠️ **{cond_sel}** tiene **{calc_preview['dias_vencidos']} días** "
                        f"de vacaciones sin registrar."
                    )

                st.divider()
                with st.form("form_vacacion_v15", clear_on_submit=True):
                    col1v, col2v = st.columns(2)
                    with col1v:
                        fi_vac = st.date_input("Fecha inicio vacaciones", value=datetime.today())
                        ff_vac = st.date_input("Fecha fin vacaciones",    value=datetime.today())
                    with col2v:
                        dias_auto = max(1, (ff_vac - fi_vac).days + 1) if ff_vac >= fi_vac else 1
                        st.metric("Días calculados automáticamente", dias_auto)
                        dias_vac  = st.number_input(
                            "Días (ajustable manualmente)",
                            min_value=1,
                            max_value=60,
                            value=min(int(dias_auto), 60)
                        )
                        reg_por_v = st.text_input("Registrado por", placeholder="Tu nombre completo")
                        obs_vac   = st.text_area("Observaciones", height=80)

                    submitted_vac = st.form_submit_button("💾 Registrar", type="primary")
                    if submitted_vac:
                        if not reg_por_v.strip():
                            st.error("⚠️ Ingresa tu nombre.")
                        elif ff_vac < fi_vac:
                            st.error("⚠️ La fecha fin no puede ser anterior a la fecha inicio.")
                        else:
                            nuevo_id_v = db.registrar_vacacion({
                                'conductor': cond_sel,
                                'fecha_inicio': fi_vac,
                                'fecha_fin': ff_vac,
                                'dias': dias_vac,
                                'anio_laboral': None,
                                'observaciones': obs_vac.strip(),
                                'registrado_por': reg_por_v.strip()
                            })
                            if nuevo_id_v:
                                st.success(
                                    f"✅ Vacaciones registradas para **{cond_sel}**: "
                                    f"{fi_vac.strftime('%d/%m/%Y')} → {ff_vac.strftime('%d/%m/%Y')} "
                                    f"({dias_vac} días)"
                                )
                                st.rerun()

                st.divider()
                st.subheader(f"Historial de vacaciones — {cond_sel}")
                df_hist_cond = df_vac_todos[df_vac_todos["conductor"] == cond_sel] \
                    if not df_vac_todos.empty else pd.DataFrame()
                if df_hist_cond.empty:
                    st.info("Sin registros de vacaciones aún.")
                else:
                    df_hist_show = df_hist_cond[["id","fecha_inicio","fecha_fin","dias","observaciones","registrado_por"]].copy()
                    # Formatear fechas DD/MM/YYYY
                    df_hist_show["fecha_inicio"] = df_hist_show["fecha_inicio"].apply(fmt_fecha)
                    df_hist_show["fecha_fin"]    = df_hist_show["fecha_fin"].apply(fmt_fecha)
                    df_hist_show.columns = ["ID","Inicio","Fin","Días","Observaciones","Registrado por"]
                    st.dataframe(df_hist_show, use_container_width=True, hide_index=True)

        # ======================== FECHA DE INGRESO ========================
        with v_tab3:
            st.subheader("⚙️ Registrar / actualizar fecha de ingreso")
            st.caption("La fecha de ingreso es la base para calcular los años laborales y los días de vacaciones.")

            with st.form("form_fecha_ingreso", clear_on_submit=True):
                col1fi, col2fi = st.columns(2)
                with col1fi:
                    cond_fi   = st.selectbox("Conductor", lista_conductores_vac, key="fi_cond")
                    fecha_ing = st.date_input("Fecha de ingreso / contratación", value=datetime.today())
                with col2fi:
                    obs_fi = st.text_area("Observaciones", height=80)

                if st.form_submit_button("💾 Guardar", type="primary"):
                    ok_fi = db.guardar_info_conductor(cond_fi, fecha_ing, obs_fi)
                    if ok_fi:
                        st.success(f"✅ Fecha de ingreso de **{cond_fi}** guardada: **{fecha_ing.strftime('%d/%m/%Y')}**")
                        st.rerun()

            st.divider()
            st.subheader("Fechas registradas y resumen rápido")
            df_info_show = db.obtener_todos_info_conductores()
            if df_info_show.empty:
                st.info("No hay fechas de ingreso registradas.")
            else:
                for _, irow in df_info_show.iterrows():
                    fi_d = pd.to_datetime(irow['fecha_ingreso']).date()
                    anios_t = round((hoy - fi_d).days / 365.25, 1)
                    calc_r  = calcular_vacaciones(irow['conductor'], fi_d, df_vac_todos, hoy)

                    if calc_r["dias_vencidos"] > 0:
                        icono = "🔴"
                    elif calc_r["anios_completos"] > 0:
                        icono = "✅"
                    else:
                        icono = "🟢"

                    st.write(
                        f"{icono} **{irow['conductor']}** — "
                        f"Ingreso: **{fi_d.strftime('%d/%m/%Y')}** — "
                        f"Antigüedad: {anios_t} años — "
                        f"Años completos: **{calc_r['anios_completos']}** — "
                        f"Días generados: **{calc_r['dias_generados']}** — "
                        f"Días usados: **{calc_r['dias_usados']}** — "
                        f"Días pendientes: **{calc_r['dias_vencidos']}** — "
                        f"Próxima vacación: **{calc_r['proxima_fecha'].strftime('%d/%m/%Y')}** "
                        f"(en {calc_r['dias_para_proxima']} días)"
                        + (f"  |  {irow['observaciones']}" if irow.get('observaciones') else "")
                    )

    # ==================== TAB 5: PRÉSTAMOS ====================
    with tab_prest:
        st.header("💰 Gestión de Préstamos a Conductores")
        lista_conductores_prest = get_lista_conductores(db)

        p_tab1, p_tab2, p_tab3 = st.tabs(["📊 Resumen y trazabilidad", "➕ Nuevo préstamo", "💳 Registrar pago/descuento"])

        with p_tab1:
            st.subheader("Estado de préstamos")
            col_fp1, col_fp2, col_fp3 = st.columns(3)
            with col_fp1: filtro_cond_p  = st.selectbox("Conductor", ["Todos"] + lista_conductores_prest, key="p_filtro_cond")
            with col_fp2: filtro_estado_p = st.selectbox("Estado", ["Todos","activo","saldado"], key="p_filtro_estado")
            with col_fp3: filtro_fecha_p  = st.date_input("Préstamos desde", value=None, key="p_filtro_fecha")

            cond_p_q = None if filtro_cond_p == "Todos" else filtro_cond_p
            est_p_q  = None if filtro_estado_p == "Todos" else filtro_estado_p
            df_prestamos_all = db.obtener_prestamos(conductor=cond_p_q, estado=est_p_q)
            df_pagos_all     = db.obtener_pagos()

            if filtro_fecha_p and not df_prestamos_all.empty:
                df_prestamos_all = df_prestamos_all[pd.to_datetime(df_prestamos_all["fecha_prestamo"]).dt.date >= filtro_fecha_p]

            if df_prestamos_all.empty:
                st.info("No se encontraron préstamos con los filtros aplicados.")
            else:
                total_prestado = int(df_prestamos_all["monto_total"].sum())
                total_pagado_g = 0; total_saldo_g = 0; activos_g = 0; saldados_g = 0
                for _, pr in df_prestamos_all.iterrows():
                    pagado, saldo = calcular_saldo_prestamo(pr['id'], pr['monto_total'], df_pagos_all)
                    total_pagado_g += pagado; total_saldo_g += saldo
                    if pr['estado'] == 'activo': activos_g += 1
                    else: saldados_g += 1

                col_pm1, col_pm2, col_pm3, col_pm4 = st.columns(4)
                col_pm1.metric("Total prestado",  f"${fmt(total_prestado)}")
                col_pm2.metric("Total pagado",    f"${fmt(total_pagado_g)}")
                col_pm3.metric("Saldo pendiente", f"${fmt(total_saldo_g)}")
                col_pm4.metric("Activos / Saldados", f"{activos_g} / {saldados_g}")

                st.divider()
                for _, pr in df_prestamos_all.iterrows():
                    pagado, saldo = calcular_saldo_prestamo(pr['id'], pr['monto_total'], df_pagos_all)
                    pct = round(pagado / pr['monto_total'] * 100) if pr['monto_total'] > 0 else 0
                    paz_salvo = pr['estado'] == 'saldado' or saldo == 0
                    icono_p = "✅" if paz_salvo else "🔴"
                    label_p = (f"{icono_p} ID {pr['id']} | {pr['conductor']} | "
                               f"${fmt(pr['monto_total'])} | Pagado: ${fmt(pagado)} | Saldo: ${fmt(saldo)} | {pct}%")

                    with st.expander(label_p):
                        col_pa, col_pb = st.columns([2, 2])
                        with col_pa:
                            st.write(f"💰 Monto: **${fmt(pr['monto_total'])} COP** | Pagado: **${fmt(pagado)}** | Saldo: **${fmt(saldo)}**")
                            if saldo > 0: st.error(f"Saldo pendiente: **${fmt(saldo)} COP**")
                            else: st.success("✅ PAZ Y SALVO")
                            st.progress(min(pct, 100) / 100, text=f"{pct}%")
                            if pr.get('motivo'): st.write(f"📝 {pr['motivo']}")
                            if not paz_salvo:
                                col_btna, col_btnb = st.columns(2)
                                with col_btna:
                                    if st.button("✅ Paz y Salvo", key=f"paz_{pr['id']}", type="primary"):
                                        db.actualizar_estado_prestamo(pr['id'], 'saldado'); st.rerun()
                                with col_btnb:
                                    if st.session_state.confirmar_eliminar_prestamo == pr['id']:
                                        st.warning("¿Eliminar?")
                                        c_s2, c_n2 = st.columns(2)
                                        with c_s2:
                                            if st.button("Sí", key=f"si_prest_{pr['id']}"):
                                                db.eliminar_prestamo(pr['id'])
                                                st.session_state.confirmar_eliminar_prestamo = None; st.rerun()
                                        with c_n2:
                                            if st.button("No", key=f"no_prest_{pr['id']}"):
                                                st.session_state.confirmar_eliminar_prestamo = None; st.rerun()
                                    else:
                                        if st.button("🗑️ Eliminar", key=f"del_prest_{pr['id']}"):
                                            st.session_state.confirmar_eliminar_prestamo = pr['id']; st.rerun()
                            else:
                                if st.button("↩️ Reabrir", key=f"reabrir_{pr['id']}"):
                                    db.actualizar_estado_prestamo(pr['id'], 'activo'); st.rerun()
                        with col_pb:
                            st.markdown("**Historial de pagos:**")
                            df_pagos_p = df_pagos_all[df_pagos_all["prestamo_id"] == pr['id']] if not df_pagos_all.empty else pd.DataFrame()
                            if df_pagos_p.empty:
                                st.info("Sin pagos registrados.")
                            else:
                                saldo_acum = int(pr['monto_total'])
                                for _, pg in df_pagos_p.sort_values("fecha_pago").iterrows():
                                    saldo_acum -= int(pg['monto_pago'])
                                    col_pgr = st.columns([3, 1])
                                    with col_pgr[0]:
                                        st.write(f"💳 {fmt_fecha(pg['fecha_pago'])} — ${fmt(pg['monto_pago'])} — Saldo: ${fmt(max(0,saldo_acum))}")
                                    with col_pgr[1]:
                                        if st.session_state.confirmar_eliminar_pago == pg['id']:
                                            c_s3, c_n3 = st.columns(2)
                                            with c_s3:
                                                if st.button("Sí", key=f"si_pago_{pg['id']}"):
                                                    db.eliminar_pago(pg['id'])
                                                    st.session_state.confirmar_eliminar_pago = None; st.rerun()
                                            with c_n3:
                                                if st.button("No", key=f"no_pago_{pg['id']}"):
                                                    st.session_state.confirmar_eliminar_pago = None; st.rerun()
                                        else:
                                            if st.button("🗑️", key=f"del_pago_{pg['id']}"):
                                                st.session_state.confirmar_eliminar_pago = pg['id']; st.rerun()

                st.divider()
                _, col_exp_p2 = st.columns([3, 1])
                with col_exp_p2:
                    df_pagos_export = db.obtener_pagos()
                    excel_p = generar_excel_prestamos(df_prestamos_all, df_pagos_export)
                    st.download_button(label="📥 Exportar a Excel", data=excel_p,
                        file_name=f"prestamos_{hora_colombia().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

        with p_tab2:
            st.subheader("Registrar nuevo préstamo")
            with st.form("form_prestamo", clear_on_submit=True):
                col1p, col2p = st.columns(2)
                with col1p:
                    cond_nuevo_p    = st.selectbox("Conductor", lista_conductores_prest, key="p_cond_nuevo")
                    fecha_prest     = st.date_input("Fecha del préstamo", value=datetime.today())
                    monto_prest_txt = st.text_input("Monto (COP)", placeholder="Ej: 500.000")
                    monto_prest     = limpiar(monto_prest_txt)
                    if monto_prest > 0: st.caption(f"💵 {fmt(monto_prest)} COP")
                with col2p:
                    motivo_prest = st.text_input("Motivo", placeholder="Ej: Urgencia médica...")
                    obs_prest    = st.text_area("Observaciones", height=80)

                if st.form_submit_button("💾 Registrar Préstamo", type="primary"):
                    if monto_prest <= 0:
                        st.error("⚠️ El monto debe ser mayor a 0.")
                    else:
                        nid_p = db.registrar_prestamo({
                            'conductor': cond_nuevo_p, 'monto_total': monto_prest,
                            'fecha_prestamo': fecha_prest, 'motivo': motivo_prest.strip(),
                            'observaciones': obs_prest.strip()
                        })
                        if nid_p:
                            st.success(f"✅ Préstamo ID {nid_p} registrado para **{cond_nuevo_p}** — ${fmt(monto_prest)} COP")
                            st.rerun()

        with p_tab3:
            st.subheader("Registrar pago / descuento")
            df_activos   = db.obtener_prestamos(estado="activo")
            df_pagos_chk = db.obtener_pagos()
            if df_activos.empty:
                st.success("✅ No hay préstamos activos.")
            else:
                opciones_prestamos = []
                for _, pr in df_activos.iterrows():
                    pagado, saldo = calcular_saldo_prestamo(pr['id'], pr['monto_total'], df_pagos_chk)
                    if saldo > 0:
                        opciones_prestamos.append({
                            "id": pr['id'],
                            "label": f"ID {pr['id']} | {pr['conductor']} | Saldo: ${fmt(saldo)}",
                            "saldo": saldo, "conductor": pr['conductor'], "monto_total": pr['monto_total']
                        })
                if not opciones_prestamos:
                    st.success("✅ Todos los préstamos activos están saldados.")
                else:
                    with st.form("form_pago", clear_on_submit=True):
                        p_idx = st.selectbox("Préstamo", range(len(opciones_prestamos)),
                            format_func=lambda i: opciones_prestamos[i]["label"], key="p_sel_pago")
                        op = opciones_prestamos[p_idx]
                        st.info(f"Saldo actual: **${fmt(op['saldo'])} COP**")
                        col1pg, col2pg = st.columns(2)
                        with col1pg:
                            fecha_pago     = st.date_input("Fecha del descuento", value=datetime.today())
                            monto_pago_txt = st.text_input("Monto del descuento (COP)")
                            monto_pago     = limpiar(monto_pago_txt)
                            if monto_pago > 0:
                                st.caption(f"Saldo restante: **${fmt(max(0, op['saldo'] - monto_pago))} COP**")
                        with col2pg:
                            reg_por_pg = st.text_input("Registrado por")
                            obs_pg     = st.text_area("Observaciones", height=80)

                        if st.form_submit_button("💳 Registrar Descuento", type="primary"):
                            if monto_pago <= 0:
                                st.error("⚠️ Monto debe ser mayor a 0.")
                            elif monto_pago > op['saldo']:
                                st.error(f"⚠️ Supera el saldo (${fmt(op['saldo'])}).")
                            elif not reg_por_pg.strip():
                                st.error("⚠️ Ingresa tu nombre.")
                            else:
                                nid_pg = db.registrar_pago({
                                    'prestamo_id': op['id'], 'monto_pago': monto_pago,
                                    'fecha_pago': fecha_pago, 'observaciones': obs_pg.strip(),
                                    'registrado_por': reg_por_pg.strip()
                                })
                                if nid_pg:
                                    nuevo_saldo = max(0, op['saldo'] - monto_pago)
                                    st.success(f"✅ Descuento registrado. Saldo: **${fmt(nuevo_saldo)} COP**")
                                    if nuevo_saldo == 0:
                                        db.actualizar_estado_prestamo(op['id'], 'saldado')
                                        st.success(f"🎉 **{op['conductor']}** — PAZ Y SALVO")
                                    st.rerun()

    # ==================== TAB 6: CLIENTES ====================
    with tab_clientes:
        st.header("🏢 Gestión de Clientes")
        st.subheader("Clientes predeterminados")
        cols = st.columns(len(CLIENTES_DEFAULT))
        for i, c_def in enumerate(CLIENTES_DEFAULT):
            with cols[i]: st.info(c_def)
        st.divider()
        with st.form("form_nuevo_cliente", clear_on_submit=True):
            nuevo_cliente = st.text_input("Nombre del cliente nuevo")
            if st.form_submit_button("➕ Agregar Cliente", type="primary"):
                if not nuevo_cliente.strip():
                    st.error("⚠️ El nombre no puede estar vacío.")
                elif nuevo_cliente.strip().upper() in [c.upper() for c in CLIENTES_DEFAULT]:
                    st.warning("⚠️ Ya existe en la lista predeterminada.")
                else:
                    ok = db.agregar_cliente(nuevo_cliente.strip())
                    if ok: st.success(f"✅ Cliente **{nuevo_cliente.strip().upper()}** agregado."); st.rerun()
                    else: st.error("❌ Ya existe o hubo un error.")
        st.divider()
        df_extras = db.obtener_clientes_extra()
        if df_extras.empty:
            st.info("No hay clientes adicionales aún.")
        else:
            for _, row in df_extras.iterrows():
                col_n, col_f, col_b = st.columns([3, 2, 1])
                with col_n: st.write(f"**{row['nombre']}**")
                with col_f: st.write(str(row['fecha_registro'])[:16])
                with col_b:
                    if st.session_state.confirmar_eliminar_cliente == row['id']:
                        c_si, c_no = st.columns(2)
                        with c_si:
                            if st.button("Sí", key=f"si_cli_{row['id']}"):
                                db.eliminar_cliente(row['id']); st.session_state.confirmar_eliminar_cliente = None; st.rerun()
                        with c_no:
                            if st.button("No", key=f"no_cli_{row['id']}"):
                                st.session_state.confirmar_eliminar_cliente = None; st.rerun()
                    else:
                        if st.button("🗑️", key=f"del_cli_{row['id']}"):
                            st.session_state.confirmar_eliminar_cliente = row['id']; st.rerun()

    # ==================== TAB 7: CONDUCTORES ====================
    with tab_conductores:
        st.header("👤 Gestión de Conductores")
        st.subheader("Conductores predeterminados")
        cols_def = st.columns(4)
        for i, c_def in enumerate(sorted(CONDUCTORES_DEFAULT)):
            with cols_def[i % 4]: st.info(c_def)
        st.divider()
        with st.form("form_nuevo_conductor", clear_on_submit=True):
            nuevo_conductor = st.text_input("Nombre del conductor nuevo")
            if st.form_submit_button("➕ Agregar Conductor", type="primary"):
                if not nuevo_conductor.strip():
                    st.error("⚠️ El nombre no puede estar vacío.")
                elif nuevo_conductor.strip().upper() in [c.upper() for c in CONDUCTORES_DEFAULT]:
                    st.warning("⚠️ Ya existe en la lista predeterminada.")
                else:
                    ok = db.agregar_conductor(nuevo_conductor.strip())
                    if ok: st.success(f"✅ **{nuevo_conductor.strip().upper()}** agregado."); st.rerun()
                    else: st.error("❌ Ya existe o hubo un error.")
        st.divider()
        df_conductores = db.obtener_conductores_extra()
        if df_conductores.empty:
            st.info("No hay conductores adicionales registrados aún.")
        else:
            for _, row in df_conductores.iterrows():
                col_nombre, col_fecha, col_edit, col_del = st.columns([3, 2, 1, 1])
                with col_nombre:
                    if st.session_state.editando_conductor_id == row['id']:
                        nombre_editado = st.text_input("Nuevo nombre", value=row['nombre'],
                            key=f"edit_input_{row['id']}", label_visibility="collapsed")
                    else:
                        st.write(f"**{row['nombre']}**")
                with col_fecha:
                    st.write(str(row['fecha_registro'])[:16])
                with col_edit:
                    if st.session_state.editando_conductor_id == row['id']:
                        if st.button("💾", key=f"save_cond_{row['id']}"):
                            if nombre_editado.strip():
                                ok = db.editar_conductor(row['id'], nombre_editado.strip())
                                if ok:
                                    st.session_state.editando_conductor_id = None; st.rerun()
                    else:
                        if st.button("✏️", key=f"edit_cond_{row['id']}"):
                            st.session_state.editando_conductor_id = row['id']; st.rerun()
                with col_del:
                    if st.session_state.editando_conductor_id == row['id']:
                        if st.button("✖", key=f"cancel_cond_{row['id']}"):
                            st.session_state.editando_conductor_id = None; st.rerun()
                    elif st.session_state.confirmar_eliminar_conductor == row['id']:
                        c_si2, c_no2 = st.columns(2)
                        with c_si2:
                            if st.button("Sí", key=f"si_cond_{row['id']}"):
                                db.eliminar_conductor(row['id']); st.session_state.confirmar_eliminar_conductor = None; st.rerun()
                        with c_no2:
                            if st.button("No", key=f"no_cond_{row['id']}"):
                                st.session_state.confirmar_eliminar_conductor = None; st.rerun()
                    else:
                        if st.button("🗑️", key=f"del_cond_{row['id']}"):
                            st.session_state.confirmar_eliminar_conductor = row['id']; st.rerun()

        st.divider()
        todos_conductores = get_lista_conductores(db)
        cols_todos = st.columns(3)
        for i, c in enumerate(todos_conductores):
            with cols_todos[i % 3]: st.write(f"• {c}")


if __name__ == "__main__":
    main()
